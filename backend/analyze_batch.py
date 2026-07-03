"""
Clasificador masivo de tickets con AI — proceso independiente (background).

Uso: python analyze_batch.py <job_id>

Lee la configuración desde STATIC_DIR/jobs/<job_id>_config.json
Escribe progreso en     STATIC_DIR/jobs/<job_id>.json
Escribe resultado en    STATIC_DIR/reports/<job_id>_<nombre>.xlsx
"""
import os
import sys
import re
import json
import time
import pyodbc
import pandas as pd
from openai import OpenAI
from datetime import datetime
from dotenv import load_dotenv

# ── entorno ──────────────────────────────────────────────────────────────────
current_dir = os.path.dirname(os.path.abspath(__file__))
env_path = os.path.join(current_dir, ".env")
load_dotenv(env_path if os.path.exists(env_path) else os.path.join(os.path.dirname(current_dir), ".env"),
            override=True)

STATIC_DIR  = os.path.join(current_dir, "static")
JOBS_DIR    = os.path.join(STATIC_DIR, "jobs")
REPORTS_DIR = os.path.join(STATIC_DIR, "reports")
os.makedirs(JOBS_DIR,    exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

DEEPSEEK_API_KEY            = os.getenv("DEEPSEEK_API_KEY", "").strip().strip('"').strip("'")
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_STORAGE_CONTAINER     = os.getenv("AZURE_STORAGE_CONTAINER", "stecnico")

BATCH_SIZE  = 50    # comentarios por llamada a DeepSeek
BATCH_DELAY = 0.5   # segundos entre llamadas (respeta rate limits)

# ── helpers de progreso ───────────────────────────────────────────────────────

def _progress_path(job_id):
    return os.path.join(JOBS_DIR, f"{job_id}.json")

def _cancel_flag_path(job_id):
    return os.path.join(JOBS_DIR, f"{job_id}_cancel.flag")

def write_progress(job_id: str, data: dict):
    """Escribe el estado del job de forma atómica."""
    path = _progress_path(job_id)
    tmp  = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)

def is_cancelled(job_id: str) -> bool:
    return os.path.exists(_cancel_flag_path(job_id))

def clear_cancel_flag(job_id: str):
    try:
        os.remove(_cancel_flag_path(job_id))
    except FileNotFoundError:
        pass

# ── base de datos ─────────────────────────────────────────────────────────────

def get_db_connection():
    driver = os.getenv("SQL_ODBC_DRIVER", "ODBC Driver 17 for SQL Server")
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={os.getenv('SQL_SERVER')};"
        f"DATABASE={os.getenv('SQL_DATABASE')};"
        f"UID={os.getenv('SQL_USER')};"
        f"PWD={os.getenv('SQL_PASSWORD')};"
        f"Encrypt=yes;TrustServerCertificate=no;"
    )
    return pyodbc.connect(conn_str)

# ── clasificación por lotes ───────────────────────────────────────────────────

CLASSIFY_SYSTEM = """Eres un clasificador preciso de tickets de servicio técnico.
Tu tarea: leer el comentario de cada técnico y decidir si el criterio indicado SE CUMPLIÓ.

REGLAS ESTRICTAS:
- "SI"         → el técnico CONFIRMA que el criterio ocurrió / fue detectado
- "NO"         → el técnico lo descarta, dice que no existe, o el tema no es relevante
- "INDEFINIDO" → comentario vacío, ilegible, o sin información suficiente para decidir

IMPORTANTE: Lee el contexto completo. Ejemplos de criterio "fuga de gas":
  "se detectó fuga en válvula" → SI
  "no había fuga de gas"       → NO   ← frase negativa = NO
  "inspección sin anomalías"   → NO
  "olor a gas en ambiente"     → SI

Responde ÚNICAMENTE con un JSON array sin texto adicional:
[{"id":"<ID>","clasificacion":"SI|NO|INDEFINIDO","justificacion":"<máx 12 palabras>"},...]"""


def classify_batch(client: OpenAI, batch: list, criteria: str,
                   col_id: str, col_text: str) -> list:
    """Envía un lote a DeepSeek y retorna lista de clasificaciones."""
    items_text = "\n".join(
        f"[{i+1}] ID={row.get(col_id,'')} | {str(row.get(col_text,'') or '(sin comentario)')[:600]}"
        for i, row in enumerate(batch)
    )
    user_msg = f"CRITERIO: {criteria}\n\nCLASIFICA ESTOS {len(batch)} COMENTARIOS:\n\n{items_text}"

    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": CLASSIFY_SYSTEM},
                    {"role": "user",   "content": user_msg}
                ],
                temperature=0.0,
                max_tokens=2048,
            )
            raw = resp.choices[0].message.content.strip()
            match = re.search(r'\[[\s\S]*\]', raw)
            if match:
                return json.loads(match.group(0))
            return []
        except Exception as e:
            if attempt < 2:
                time.sleep(2 ** attempt)
            else:
                print(f"[batch] Error tras 3 intentos: {e}", file=sys.stderr)
    return []

# ── excel con formato ─────────────────────────────────────────────────────────

def write_excel(df: pd.DataFrame, filepath: str, classif_col: str, justif_col: str):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SI_FILL     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    NO_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    INDEF_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    SI_FONT     = Font(bold=True, color="375623")
    NO_FONT     = Font(bold=True, color="9C0006")
    INDEF_FONT  = Font(bold=True, color="7D6608")

    cols = list(df.columns)
    classif_idx = cols.index(classif_col) + 1 if classif_col in cols else None
    justif_idx  = cols.index(justif_col)  + 1 if justif_col  in cols else None

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Analisis")
        ws = writer.sheets["Analisis"]

        # Encabezados
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Colorear columna de clasificación
        if classif_idx:
            for row_i in range(2, len(df) + 2):
                cell = ws.cell(row=row_i, column=classif_idx)
                v = str(cell.value or "")
                if v == "SI":
                    cell.fill, cell.font = SI_FILL, SI_FONT
                elif v == "NO":
                    cell.fill, cell.font = NO_FILL, NO_FONT
                elif v == "INDEFINIDO":
                    cell.fill, cell.font = INDEF_FILL, INDEF_FONT

        # Anchos de columna (limitado a 50 k filas para no ser lento)
        if len(df) <= 50_000:
            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[letter].width = min(max_len + 4, 60)
        else:
            # Anchos fijos para datasets grandes
            for c_idx, col_name in enumerate(cols, 1):
                letter = get_column_letter(c_idx)
                if col_name == classif_col:
                    ws.column_dimensions[letter].width = 15
                elif col_name == justif_col:
                    ws.column_dimensions[letter].width = 55
                else:
                    ws.column_dimensions[letter].width = 20

        if justif_idx:
            ws.column_dimensions[get_column_letter(justif_idx)].width = 55

        ws.freeze_panes = "A2"

# ── subida a azure ────────────────────────────────────────────────────────────

def upload_to_azure(filepath: str, filename: str) -> str:
    if not AZURE_STORAGE_CONNECTION_STRING:
        return f"/static/reports/{os.path.basename(filepath)}"
    try:
        from azure.storage.blob import BlobServiceClient, ContentSettings
        blob_name = f"generated/reports/{filename}"
        bsc = BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
        bc  = bsc.get_blob_client(container=AZURE_STORAGE_CONTAINER, blob=blob_name)
        with open(filepath, "rb") as data:
            bc.upload_blob(data, overwrite=True, content_settings=ContentSettings(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ))
        try:
            os.remove(filepath)
        except Exception:
            pass
        return bc.url
    except Exception as e:
        print(f"[azure] Error subiendo: {e}", file=sys.stderr)
        return f"/static/reports/{os.path.basename(filepath)}"

# ── proceso principal ─────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python analyze_batch.py <job_id>", file=sys.stderr)
        sys.exit(1)

    job_id = sys.argv[1]
    config_path = os.path.join(JOBS_DIR, f"{job_id}_config.json")

    if not os.path.exists(config_path):
        print(f"[ERROR] Config no encontrada: {config_path}", file=sys.stderr)
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    sql_query    = cfg["sql_query"]
    criteria     = cfg["criteria"]
    col_id       = cfg["columna_id"]
    col_text     = cfg["columna_texto"]
    nombre       = cfg["nombre_reporte"]

    def upd(msg, procesados=0, total=0, si=0, no=0, indef=0, pct=0, status="running", url=None):
        write_progress(job_id, {
            "status": status, "message": msg,
            "procesados": procesados, "total": total,
            "si": si, "no": no, "indefinido": indef,
            "pct": pct, "download_url": url,
            "updated_at": datetime.now().isoformat()
        })

    try:
        # 1. Conectar y ejecutar SQL sin límite de filas
        upd("Conectando a la base de datos...")
        conn = get_db_connection()
        upd("Ejecutando consulta SQL...")
        df = pd.read_sql(sql_query, conn)
        conn.close()

        total = len(df)
        if total == 0:
            upd("La consulta no devolvió datos.", status="completed", pct=100)
            return

        upd(f"Iniciando clasificación de {total:,} tickets...", total=total)

        # 2. Clasificar en lotes
        client  = OpenAI(api_key=DEEPSEEK_API_KEY, base_url="https://api.deepseek.com")
        records = df.to_dict(orient="records")
        results = []
        si_n = no_n = indef_n = 0

        for start in range(0, total, BATCH_SIZE):
            if is_cancelled(job_id):
                clear_cancel_flag(job_id)
                upd(f"Cancelado por el usuario. Procesados: {start:,} de {total:,}.",
                    procesados=start, total=total, si=si_n, no=no_n, indef=indef_n,
                    pct=round(start/total*100, 1), status="cancelled")
                return

            batch        = records[start:start + BATCH_SIZE]
            classifs     = classify_batch(client, batch, criteria, col_id, col_text)
            classif_map  = {str(c.get("id", "")): c for c in classifs}

            for row in batch:
                rid = str(row.get(col_id, ""))
                c   = classif_map.get(rid, {})
                row["Clasificacion"]    = c.get("clasificacion", "INDEFINIDO")
                row["Justificacion_AI"] = c.get("justificacion", "")
                results.append(row)

                if row["Clasificacion"] == "SI":
                    si_n += 1
                elif row["Clasificacion"] == "NO":
                    no_n += 1
                else:
                    indef_n += 1

            done = min(start + BATCH_SIZE, total)
            pct  = round(done / total * 100, 1)
            upd(f"Procesando… {done:,} / {total:,} tickets ({pct}%)",
                procesados=done, total=total, si=si_n, no=no_n, indef=indef_n, pct=pct)

            time.sleep(BATCH_DELAY)

        # 3. Construir DataFrame resultado
        result_df = pd.DataFrame(results)

        # Ordenar: SI primero, luego INDEFINIDO, luego NO
        sort_map = {"SI": 0, "INDEFINIDO": 1, "NO": 2}
        result_df["_sort"] = result_df["Clasificacion"].map(sort_map).fillna(3)
        result_df = result_df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

        # 4. Generar Excel
        upd("Generando archivo Excel…", procesados=total, total=total,
            si=si_n, no=no_n, indef=indef_n, pct=100)

        ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r"[^\w\-]", "_", nombre).lower()
        filename  = f"{safe_name}_{ts}.xlsx"
        filepath  = os.path.join(REPORTS_DIR, filename)

        write_excel(result_df, filepath, "Clasificacion", "Justificacion_AI")

        # 5. Subir a Azure
        download_url = upload_to_azure(filepath, filename)

        # 6. Estado final
        upd(
            f"✅ Análisis completado. {si_n:,} positivos de {total:,} tickets analizados.",
            procesados=total, total=total, si=si_n, no=no_n, indef=indef_n,
            pct=100, status="completed", url=download_url
        )

    except Exception as e:
        print(f"[FATAL] {e}", file=sys.stderr)
        try:
            upd(f"Error fatal: {str(e)}", status="error")
        except Exception:
            pass


if __name__ == "__main__":
    main()
