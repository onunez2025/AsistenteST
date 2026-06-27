import os
import sys
import re
import uuid
import subprocess
import logging
import pyodbc
import pandas as pd
import plotly.express as px
from datetime import datetime
from dotenv import load_dotenv
from fastmcp import FastMCP
from azure.storage.blob import BlobServiceClient, ContentSettings
from typing import Optional
import json
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Setup logging to stderr because stdout is used for MCP stdio protocol communication
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stderr
)
logger = logging.getLogger("mcp-db")

# Load environment variables dynamically using absolute paths
current_dir = os.path.dirname(os.path.abspath(__file__))
env_in_backend = os.path.join(current_dir, ".env")
env_in_root = os.path.join(os.path.dirname(current_dir), ".env")

if os.path.exists(env_in_backend):
    load_dotenv(env_in_backend, override=True)
elif os.path.exists(env_in_root):
    load_dotenv(env_in_root, override=True)
else:
    load_dotenv(override=True)

# SQL Azure Config
SQL_SERVER = os.getenv("SQL_SERVER")
SQL_DATABASE = os.getenv("SQL_DATABASE")
SQL_USER = os.getenv("SQL_USER")
SQL_PASSWORD = os.getenv("SQL_PASSWORD")

# Azure Blob Storage Config
AZURE_STORAGE_CONNECTION_STRING = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZURE_STORAGE_CONTAINER = os.getenv("AZURE_STORAGE_CONTAINER", "stecnico")

def upload_file_to_azure_blob(local_filepath: str, blob_name: str, content_type: str) -> tuple[Optional[str], Optional[str]]:
    """Uploads a local file to Azure Blob Storage and returns a tuple (url, error_message)."""
    if not AZURE_STORAGE_CONNECTION_STRING:
        logger.warning("AZURE_STORAGE_CONNECTION_STRING no configurada. Retornando local path.")
        return None, "AZURE_STORAGE_CONNECTION_STRING no configurada en las variables de entorno."
    try:
        blob_service_client = BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
        blob_client = blob_service_client.get_blob_client(container=AZURE_STORAGE_CONTAINER, blob=blob_name)
        
        # Upload the file
        with open(local_filepath, "rb") as data:
            blob_client.upload_blob(data, overwrite=True, content_settings=ContentSettings(content_type=content_type))
        
        # URL of the uploaded blob
        url = blob_client.url
        logger.info(f"Archivo subido exitosamente a Azure Blob Storage: {url}")
        return url, None
    except Exception as e:
        err_msg = str(e)
        logger.error(f"Error subiendo archivo a Azure Blob Storage: {err_msg}")
        return None, err_msg


# Define directories relative to this file's location to ensure correctness
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR  = os.path.join(BACKEND_DIR, "static")
REPORTS_DIR = os.path.join(STATIC_DIR, "reports")
CHARTS_DIR  = os.path.join(STATIC_DIR, "charts")
JOBS_DIR    = os.path.join(STATIC_DIR, "jobs")

os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(CHARTS_DIR,  exist_ok=True)
os.makedirs(JOBS_DIR,    exist_ok=True)

# Initialize FastMCP Server
mcp = FastMCP("Azure SQL Database Server")

def get_db_connection():
    """Establishes connection to Azure SQL Database."""
    conn_str = (
        f"DRIVER={{ODBC Driver 17 for SQL Server}};"
        f"SERVER={SQL_SERVER};"
        f"DATABASE={SQL_DATABASE};"
        f"UID={SQL_USER};"
        f"PWD={SQL_PASSWORD};"
        f"Encrypt=yes;"
        f"TrustServerCertificate=no;"
    )
    return pyodbc.connect(conn_str)

def is_query_safe(query: str) -> bool:
    """Verifies that the query is read-only (SELECT or WITH)."""
    clean_query = query.strip().upper()
    if not (clean_query.startswith("SELECT") or clean_query.startswith("WITH")):
        return False
    dangerous_keywords = ["INSERT", "UPDATE", "DELETE", "DROP", "ALTER", "TRUNCATE", "CREATE", "EXEC", "EXECUTE", "GRANT", "REVOKE"]
    for kw in dangerous_keywords:
        pattern = r"\b" + re.escape(kw) + r"\b"
        if re.search(pattern, clean_query):
            return False
    return True

@mcp.tool()
def ejecutar_consulta_sql(sql_query: str) -> str:
    """
    Ejecuta una consulta SQL de solo lectura (SELECT) en la base de datos de Azure SQL 
    y devuelve los resultados en formato JSON. Esta herramienta sirve para obtener datos,
    hacer sumatorias, promedios, listados de técnicos, estados de órdenes o detalles
    específicos guardados en la tabla 'APPGAC.ServiciosViewSQL'.
    
    Args:
        sql_query: La consulta SQL SELECT a ejecutar. Debe ser válida para Microsoft SQL Server.
    """
    logger.info(f"[MCP DB] ejecutar_consulta_sql: {sql_query}")
    
    if not is_query_safe(sql_query):
        return "Error: Solo se permiten consultas de lectura (SELECT o WITH). No se permiten operaciones de modificación."
    
    try:
        conn = get_db_connection()
        df = pd.read_sql(sql_query, conn)
        conn.close()
        
        limit = 500
        total_rows = len(df)
        if total_rows > limit:
            df_limited = df.head(limit)
            result_json = df_limited.to_json(orient="records", date_format="iso")
            return f"Resultados (Primeros {limit} de {total_rows} filas totales):\n{result_json}"
        else:
            return df.to_json(orient="records", date_format="iso")
            
    except Exception as e:
        logger.error(f"Error en ejecutar_consulta_sql: {e}")
        return f"Error al ejecutar la consulta SQL en el servidor MCP: {str(e)}"

@mcp.tool()
def generar_reporte_excel(sql_query: str, nombre_reporte: str) -> str:
    """
    Ejecuta una consulta SQL y genera un archivo Excel (.xlsx) descargable con los datos.
    Debe llamarse cuando el usuario pida explícitamente descargar, exportar, o crear un reporte
    en Excel/CSV. Devuelve el enlace de descarga del archivo.
    
    Args:
        sql_query: Consulta SQL SELECT para obtener los datos del reporte. NUNCA uses SELECT *.
        nombre_reporte: Nombre de archivo para el reporte (ej. 'servicios_mayo_2026').
    """
    logger.info(f"[MCP DB] generar_reporte_excel: {nombre_reporte}")
    if not is_query_safe(sql_query):
        return "Error: Solo se permiten consultas SELECT de lectura para generar reportes."
        
    try:
        conn = get_db_connection()
        df = pd.read_sql(sql_query, conn)
        conn.close()
        
        if df.empty:
            return "La consulta no devolvió datos. No se generó el archivo Excel."
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r"[^\w\-]", "_", nombre_reporte).lower()
        filename = f"{safe_name}_{timestamp}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Reporte")
            ws = writer.sheets["Reporte"]

            # Estilo de encabezados
            header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            ws.row_dimensions[1].height = 22

            # Auto-ancho de columnas
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=8)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

            # Fijar primera fila
            ws.freeze_panes = "A2"
        
        url = f"/static/reports/{filename}"
        return f"Reporte Excel generado con éxito. Descárgalo aquí: [Descargar Reporte Excel]({url})"

        
    except Exception as e:
        logger.error(f"Error generando Excel en MCP: {e}")
        return f"Error al generar el reporte Excel: {str(e)}"

@mcp.tool()
def generar_grafico(sql_query: str, tipo_grafico: str, columna_x: str, columna_y: str, titulo: str) -> str:
    """
    Ejecuta una consulta SQL, analiza los datos y genera un gráfico interactivo en formato HTML
    (usando Plotly). Devuelve el enlace de visualización para incrustar el gráfico en el chat.

    Args:
        sql_query: Consulta SQL SELECT para obtener los datos del gráfico.
        tipo_grafico: Tipo de gráfico: 'bar' (barras verticales), 'bar_h' (barras horizontales — ideal para rankings de técnicos/CAS), 'line' (línea de tendencia), 'pie' (torta/porcentajes), 'scatter' (dispersión), 'funnel' (embudo — para pipelines), 'histogram' (distribución de frecuencia de columna_x).
        columna_x: Columna para el eje X (categorías). En bar_h: es el eje Y (categorías). En histogram: columna a analizar.
        columna_y: Columna para el eje Y (valores numéricos). En bar_h: es el eje X (valores). En histogram: se ignora.
        titulo: Título del gráfico.
    """
    logger.info(f"[MCP DB] generar_grafico: {tipo_grafico} | {titulo}")
    if not is_query_safe(sql_query):
        return "Error: Solo se permiten consultas SELECT de lectura para generar gráficos."
        
    try:
        conn = get_db_connection()
        df = pd.read_sql(sql_query, conn)
        conn.close()
        
        if df.empty:
            return "La consulta no devolvió datos. No se pudo generar el gráfico."
            
        if columna_x not in df.columns or columna_y not in df.columns:
            return f"Error: Las columnas '{columna_x}' o '{columna_y}' no existen. Disponibles: {list(df.columns)}"
            
        fig = None
        tipo_grafico = tipo_grafico.lower().strip()

        if tipo_grafico == "bar":
            fig = px.bar(df, x=columna_x, y=columna_y, title=titulo, template="plotly_dark", text_auto=True)
        elif tipo_grafico == "bar_h":
            fig = px.bar(df, y=columna_x, x=columna_y, title=titulo, template="plotly_dark",
                         orientation="h", text_auto=True)
        elif tipo_grafico == "line":
            fig = px.line(df, x=columna_x, y=columna_y, title=titulo, template="plotly_dark", markers=True)
        elif tipo_grafico == "pie":
            fig = px.pie(df, names=columna_x, values=columna_y, title=titulo, template="plotly_dark")
        elif tipo_grafico == "scatter":
            fig = px.scatter(df, x=columna_x, y=columna_y, title=titulo, template="plotly_dark")
        elif tipo_grafico == "funnel":
            fig = px.funnel(df, x=columna_y, y=columna_x, title=titulo, template="plotly_dark")
        elif tipo_grafico == "histogram":
            fig = px.histogram(df, x=columna_x, title=titulo, template="plotly_dark")
        else:
            fig = px.bar(df, x=columna_x, y=columna_y, title=titulo, template="plotly_dark", text_auto=True)
            
        fig.update_layout(
            paper_bgcolor="rgba(24, 24, 28, 0.95)",
            plot_bgcolor="rgba(0, 0, 0, 0)",
            title_font=dict(size=18, family="Outfit, sans-serif", color="#ffffff"),
            font=dict(family="Inter, sans-serif", color="#a1a1aa"),
        )
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = re.sub(r"[^\w\-]", "_", titulo).lower()
        filename = f"chart_{safe_title}_{timestamp}.html"
        filepath = os.path.join(CHARTS_DIR, filename)
        
        fig.write_html(filepath, include_plotlyjs="cdn", full_html=True)

        url = f"/static/charts/{filename}"
        return f"Gráfico interactivo generado con éxito. [EmbedChart:{url}]"

        
    except Exception as e:
        logger.error(f"Error generando gráfico en MCP: {e}")
        return f"Error al generar el gráfico: {str(e)}"

SHARED_KNOWLEDGE_FILE = os.path.join(BACKEND_DIR, "shared_knowledge.json")

@mcp.tool()
def guardar_regla_negocio(tema: str, explicacion: str, consulta_sql: Optional[str] = None) -> str:
    """
    Guarda una regla de negocio, fórmula, cruce de tablas, método de cálculo de indicador 
    o instrucción enseñada por el usuario para que el asistente pueda recordarla y usarla 
    en el futuro en cualquier conversación.
    
    Args:
        tema: El nombre del indicador, regla o tema (ej. 'Cálculo de NPS', 'Join de técnicos propios').
        explicacion: La explicación detallada de la regla, fórmula o lógica.
        consulta_sql: Un ejemplo de consulta SQL opcional que represente la regla.
    """
    logger.info(f"[MCP DB] guardar_regla_negocio: {tema}")
    try:
        data = []
        if os.path.exists(SHARED_KNOWLEDGE_FILE):
            try:
                with open(SHARED_KNOWLEDGE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                data = []
        
        nueva_regla = {
            "tema": tema,
            "explicacion": explicacion,
            "consulta_sql": consulta_sql,
            "fecha_creacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Avoid duplicate topics (update if already exists)
        exists = False
        for i, item in enumerate(data):
            if item["tema"].lower().strip() == tema.lower().strip():
                data[i] = nueva_regla
                exists = True
                break
        
        if not exists:
            data.append(nueva_regla)
            
        with open(SHARED_KNOWLEDGE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        return f"Regla de negocio sobre '{tema}' guardada exitosamente en la memoria compartida."
    except Exception as e:
        logger.error(f"Error en guardar_regla_negocio: {e}")
        return f"Error al guardar la regla de negocio: {str(e)}"

@mcp.tool()
def buscar_reglas_negocio(termino_busqueda: Optional[str] = None) -> str:
    """
    Busca e investiga en la memoria compartida las reglas de negocio, fórmulas o lógicas 
    que los usuarios han enseñado previamente al asistente. Úsala siempre que te pregunten 
    por un cálculo, indicador personalizado o cruce de tablas del cual no recuerdes la lógica exacta.
    
    Args:
        termino_busqueda: Término de búsqueda opcional para filtrar los temas. Si es vacío, devuelve todas las reglas.
    """
    logger.info(f"[MCP DB] buscar_reglas_negocio: {termino_busqueda}")
    try:
        if not os.path.exists(SHARED_KNOWLEDGE_FILE):
            return "No hay ninguna regla de negocio guardada en la memoria compartida aún."
            
        with open(SHARED_KNOWLEDGE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            
        if not data:
            return "No hay ninguna regla de negocio registrada en la memoria."
            
        if termino_busqueda:
            term = termino_busqueda.lower().strip()
            filtrados = [
                item for item in data 
                if term in item["tema"].lower() or term in item["explicacion"].lower() or (item["consulta_sql"] and term in item["consulta_sql"].lower())
            ]
            if not filtrados:
                return f"No se encontraron reglas de negocio asociadas al término '{termino_busqueda}'."
            return json.dumps(filtrados, ensure_ascii=False, indent=2)
        else:
            return json.dumps(data, ensure_ascii=False, indent=2)
            
    except Exception as e:
        logger.error(f"Error en buscar_reglas_negocio: {e}")
        return f"Error al buscar reglas de negocio: {str(e)}"

@mcp.tool()
def iniciar_analisis_masivo(
    sql_query: str,
    criterio_clasificacion: str,
    columna_id: str,
    columna_texto: str,
    nombre_reporte: str
) -> str:
    """
    Lanza un análisis masivo de tickets en SEGUNDO PLANO: ejecuta una consulta SQL SIN límite
    de filas, clasifica cada comentario usando AI (DeepSeek) según el criterio indicado, y genera
    un Excel descargable con columnas 'Clasificacion' (SI/NO/INDEFINIDO) y 'Justificacion_AI'.

    Úsala cuando el usuario necesite clasificar grandes volúmenes de tickets según un criterio
    semántico que requiere leer y entender el comentario del técnico (no solo buscar palabras clave).
    Ejemplos: "tickets con fuga de gas confirmada", "servicios donde el equipo fue reemplazado",
    "visitas donde el cliente estaba ausente".

    El proceso corre en background. Devuelve un job_id para consultar el progreso
    con verificar_estado_analisis(job_id). NO bloquea la conversación.

    Args:
        sql_query: SQL SELECT sin límite de filas. DEBE incluir la columna ID y la columna de texto.
                   Recomendado: pre-filtrar por fechas y excluir comentarios nulos para reducir volumen.
                   Ejemplo: SELECT Ticket, FechaVisita, NombreTecnico, ComentarioTecnico
                            FROM [APPGAC].[ServiciosViewSQL]
                            WHERE YEAR(FechaVisita) = 2026 AND ComentarioTecnico IS NOT NULL
                            AND LEN(LTRIM(ComentarioTecnico)) > 5
        criterio_clasificacion: Descripción clara y precisa del criterio para responder SI o NO.
                                  Ser específico mejora la precisión. Ejemplo:
                                  "El técnico detectó y confirmó una fuga de gas real. NO cuenta: inspecciones
                                  preventivas sin fuga, ni comentarios que digan que no había fuga."
        columna_id: Nombre exacto de la columna que identifica cada ticket (ej: 'Ticket').
        columna_texto: Nombre exacto de la columna con el texto a clasificar (ej: 'ComentarioTecnico').
        nombre_reporte: Nombre del archivo Excel resultante (ej: 'fugas_gas_2026').
    """
    job_id = uuid.uuid4().hex[:10]

    config = {
        "job_id":        job_id,
        "sql_query":     sql_query,
        "criteria":      criterio_clasificacion,
        "columna_id":    columna_id,
        "columna_texto": columna_texto,
        "nombre_reporte": nombre_reporte,
        "created_at":    datetime.now().isoformat()
    }

    config_path = os.path.join(JOBS_DIR, f"{job_id}_config.json")
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False)

    # Crear progress inicial para que verificar_estado_analisis no devuelva "no encontrado"
    progress_path = os.path.join(JOBS_DIR, f"{job_id}.json")
    with open(progress_path, "w", encoding="utf-8") as f:
        json.dump({
            "status": "starting", "message": "Iniciando análisis...",
            "procesados": 0, "total": 0, "si": 0, "no": 0, "indefinido": 0,
            "pct": 0, "download_url": None, "updated_at": datetime.now().isoformat()
        }, f, ensure_ascii=False)

    script_path = os.path.join(BACKEND_DIR, "analyze_batch.py")
    kwargs = {
        "stdout": subprocess.DEVNULL,
        "stderr": subprocess.DEVNULL,
        "cwd": BACKEND_DIR,
    }
    if sys.platform == "win32":
        kwargs["creationflags"] = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP
    else:
        kwargs["close_fds"] = True

    subprocess.Popen([sys.executable, script_path, job_id], **kwargs)

    tiempo_est = "desconocido (depende del volumen)"
    return (
        f"✅ Análisis masivo iniciado en segundo plano.\n\n"
        f"**Job ID:** `{job_id}`\n"
        f"**Criterio:** {criterio_clasificacion}\n"
        f"**Columna analizada:** {columna_texto}\n"
        f"**Reporte:** {nombre_reporte}\n\n"
        f"Velocidad estimada: ~6,000 tickets/minuto.\n"
        f"- 10,000 tickets → ~2 min\n"
        f"- 100,000 tickets → ~17 min\n"
        f"- 700,000 tickets → ~2 horas\n\n"
        f"Consulta el progreso con: `verificar_estado_analisis('{job_id}')`"
    )


@mcp.tool()
def verificar_estado_analisis(job_id: str) -> str:
    """
    Consulta el progreso de un análisis masivo lanzado con iniciar_analisis_masivo.
    Devuelve el porcentaje completado, conteos parciales y el enlace de descarga
    cuando el análisis termina. Úsala periódicamente para informar al usuario del avance.

    Args:
        job_id: ID devuelto por iniciar_analisis_masivo.
    """
    path = os.path.join(JOBS_DIR, f"{job_id}.json")
    if not os.path.exists(path):
        return f"No se encontró el análisis con ID '{job_id}'. Verifica que el ID sea correcto."

    try:
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
    except Exception as e:
        return f"Error leyendo estado del análisis: {e}"

    status      = d.get("status", "unknown")
    procesados  = d.get("procesados", 0)
    total       = d.get("total", 0)
    si_n        = d.get("si", 0)
    no_n        = d.get("no", 0)
    indef_n     = d.get("indefinido", 0)
    pct         = d.get("pct", 0)
    msg         = d.get("message", "")
    url         = d.get("download_url")

    if status in ("starting", "running"):
        barra = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
        return (
            f"⏳ **Análisis en progreso** — {pct}%\n"
            f"`{barra}`\n\n"
            f"Procesados: {procesados:,} de {total:,} tickets\n"
            f"✅ SI: {si_n:,}  ❌ NO: {no_n:,}  ❓ Indefinidos: {indef_n:,}\n\n"
            f"_{msg}_"
        )
    elif status == "completed":
        return (
            f"✅ **Análisis completado.**\n\n"
            f"| Resultado | Tickets |\n"
            f"|---|---|\n"
            f"| Total analizados | {total:,} |\n"
            f"| ✅ Positivos (SI) | {si_n:,} |\n"
            f"| ❌ Negativos (NO) | {no_n:,} |\n"
            f"| ❓ Indefinidos | {indef_n:,} |\n\n"
            f"[Descargar Reporte Excel]({url})"
        )
    elif status == "cancelled":
        return (
            f"🚫 Análisis cancelado.\n"
            f"Se procesaron {procesados:,} de {total:,} tickets antes de cancelar.\n"
            f"SI: {si_n:,} | NO: {no_n:,} | Indefinidos: {indef_n:,}"
        )
    elif status == "error":
        return f"🔴 **Error en el análisis:** {msg}"
    else:
        return f"Estado desconocido: {status} — {msg}"


@mcp.tool()
def cancelar_analisis(job_id: str) -> str:
    """
    Cancela un análisis masivo en curso lanzado con iniciar_analisis_masivo.
    El proceso se detiene al finalizar el lote actual (puede tardar hasta 30 segundos).
    Los tickets ya procesados se conservan en el progreso.

    Args:
        job_id: ID del análisis a cancelar.
    """
    progress_path = os.path.join(JOBS_DIR, f"{job_id}.json")
    cancel_flag   = os.path.join(JOBS_DIR, f"{job_id}_cancel.flag")

    if not os.path.exists(progress_path):
        return f"No se encontró el análisis con ID '{job_id}'."

    with open(progress_path, "r", encoding="utf-8") as f:
        d = json.load(f)

    if d.get("status") not in ("starting", "running"):
        return f"El análisis '{job_id}' no está en ejecución (estado: {d.get('status')})."

    with open(cancel_flag, "w") as f:
        f.write("cancel")

    return (
        f"✅ Señal de cancelación enviada al análisis `{job_id}`.\n"
        f"El proceso se detendrá al finalizar el lote actual (máx. ~30 segundos)."
    )


if __name__ == "__main__":
    mcp.run()
